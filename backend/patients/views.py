from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Patient, SymptomLog
from openpyxl import Workbook
import datetime
from django.views.decorators.csrf import csrf_exempt

def symptom_log_view(request):
    symptom_logs = SymptomLog.objects.all().order_by('-date_logged')
    return render(request, 'symptoms/symptom_log.html', {
        'symptom_logs': symptom_logs,
        'gender_choices': SymptomLog.GENDER_CHOICES,
    })

@csrf_exempt
def add_symptom(request):
    if request.method == 'POST':
        SymptomLog.objects.create(
            age=request.POST.get('age'),
            gender=request.POST.get('gender'),
            symptom=request.POST.get('symptom'),
        )
        return redirect('symptom_log')

    return redirect('symptom_log')


def edit_symptom(request, log_id):
    log = get_object_or_404(SymptomLog, id=log_id)

    if request.method == 'POST':
        log.age = request.POST.get('age')
        log.gender = request.POST.get('gender')
        log.symptom = request.POST.get('symptom')
        log.save()
        return redirect('symptom_log')

    symptom_logs = SymptomLog.objects.all().order_by('-date_logged')
    return render(request, 'symptoms/symptom_log.html', {
        'symptom_logs': symptom_logs,
        'edit_log': log,
        'gender_choices': SymptomLog.GENDER_CHOICES,
    })


def delete_symptom(request, log_id):
    log = get_object_or_404(SymptomLog, id=log_id)
    log.delete()
    return redirect('symptom_log')

@login_required
def symptom_logs_table(request):
    symptom_logs = SymptomLog.objects.all().order_by('-patient__date_logged')
    return render(request, 'patients/symptom_logs.html', {
        'symptom_logs': symptom_logs,
    })

@login_required
def add_symptom_log(request):
    if request.method == 'POST':
        try:
            patient_name = request.POST.get('patient_name')
            age = request.POST.get('age')
            sex = request.POST.get('sex')
            symptom = request.POST.get('symptom')
            notes = request.POST.get('notes', '')
            
            # Validate required fields
            if not patient_name:
                return JsonResponse({'success': False, 'error': 'Patient name is required'})
            if not age:
                return JsonResponse({'success': False, 'error': 'Age is required'})
            if not sex:
                return JsonResponse({'success': False, 'error': 'Sex is required'})
            if not symptom:
                return JsonResponse({'success': False, 'error': 'Symptom is required'})
            
            # Create or get patient
            patient, created = Patient.objects.get_or_create(
                name=patient_name,
                defaults={'age': int(age), 'sex': sex}
            )
            
            # Create symptom log
            symptom_log = SymptomLog.objects.create(
                patient=patient,
                symptom=symptom,
                notes=notes
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Symptom log added successfully!',
                'log_data': {
                    'id': symptom_log.id,
                    'patient_name': patient.name,
                    'age': patient.age,
                    'sex': patient.sex,
                    'symptom': symptom_log.symptom,
                    'notes': symptom_log.notes,
                    'date_logged': patient.date_logged.strftime('%Y-%m-%d')
                }
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': 'Invalid age - must be a number'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error creating log: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required 
def edit_symptom_log(request, log_id):
    symptom_log = get_object_or_404(SymptomLog, id=log_id)
    
    if request.method == 'POST':
        try:
            patient_name = request.POST.get('patient_name')
            age = request.POST.get('age')
            sex = request.POST.get('sex')
            symptom = request.POST.get('symptom')
            notes = request.POST.get('notes', '')
            
            # Validate required fields
            if not patient_name:
                return JsonResponse({'success': False, 'error': 'Patient name is required'})
            if not age:
                return JsonResponse({'success': False, 'error': 'Age is required'})
            if not sex:
                return JsonResponse({'success': False, 'error': 'Sex is required'})
            if not symptom:
                return JsonResponse({'success': False, 'error': 'Symptom is required'})
            
            # Update patient info
            patient = symptom_log.patient
            patient.name = patient_name
            patient.age = int(age)
            patient.sex = sex
            patient.save()
            
            # Update symptom log
            symptom_log.symptom = symptom
            symptom_log.notes = notes
            symptom_log.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Symptom log updated successfully!',
                'log_data': {
                    'id': symptom_log.id,
                    'patient_name': patient.name,
                    'age': patient.age,
                    'sex': patient.sex,
                    'symptom': symptom_log.symptom,
                    'notes': symptom_log.notes,
                    'date_logged': patient.date_logged.strftime('%Y-%m-%d')
                }
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': 'Invalid age - must be a number'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error updating log: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def delete_symptom_log(request, log_id):
    if request.method == 'POST':
        try:
            symptom_log = get_object_or_404(SymptomLog, id=log_id)
            symptom_log.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Symptom log deleted successfully!'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error deleting log: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def export_symptom_logs_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=symptom_logs.xlsx'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Symptom Logs'
    
    # Add headers
    headers = ['Patient Name', 'Age', 'Sex', 'Symptom', 'Notes', 'Date Logged']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Add data
    symptom_logs = SymptomLog.objects.all().order_by('-patient__date_logged')
    for row_num, log in enumerate(symptom_logs, 2):
        worksheet.cell(row=row_num, column=1, value=log.patient.name)
        worksheet.cell(row=row_num, column=2, value=log.patient.age)
        worksheet.cell(row=row_num, column=3, value=log.patient.sex)
        worksheet.cell(row=row_num, column=4, value=log.symptom)
        worksheet.cell(row=row_num, column=5, value=log.notes)
        worksheet.cell(row=row_num, column=6, value=log.patient.date_logged.strftime('%Y-%m-%d'))
    
    workbook.save(response)
    return response